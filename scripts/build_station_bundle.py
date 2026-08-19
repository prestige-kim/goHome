#!/usr/bin/env python3

"""Build the GoHome station bundle from official KRIC and Seoul source files."""

from __future__ import annotations

import argparse
import json
import math
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_KRIC_PATH = PROJECT_ROOT / "DataSources/raw/kric_metro_stations_20260630.xlsx"
DEFAULT_SEOUL_PATH = PROJECT_ROOT / "DataSources/raw/seoul_realtime_arrival_stations_20260804.xlsx"
DEFAULT_GTX_PATH = PROJECT_ROOT / "DataSources/raw/kric_gtx_a_stations_20240715.xlsx"
DEFAULT_OUTPUT_PATH = PROJECT_ROOT / "GoHome/Resources/stations.seed.json"
DEFAULT_REPORT_PATH = PROJECT_ROOT / "DataSources/station_build_report.json"

CLUSTER_DISTANCE_METERS = 850.0

# Names alone cannot always identify one physical station. These pairs share a
# name but are not transfer stations even though some entrances are relatively
# close. Keep each line as a separate selectable station.
DISTINCT_SAME_NAME_LINES = {
    "신촌": ({"2호선"}, {"경의중앙선"}),
    "양평": ({"5호선"}, {"경의중앙선"}),
}

# The current KRIC workbook places the 4호선 이촌 coordinate about 900 m from
# the connected 경의중앙선 platform. Merge the official same-name records and
# use the KORAIL coordinate from that workbook for proximity calculations.
FORCED_TRANSFER_NAMES = {"이촌"}
PHYSICAL_STATION_COORDINATE_OVERRIDES = {
    "이촌": (37.522476, 126.973816),
}

LINE_ALIASES: dict[str, tuple[str, ...]] = {
    "1호선": ("1호선", "경부선", "경원선", "경인선", "장항선"),
    "2호선": ("2호선",),
    "3호선": ("3호선", "일산선"),
    "4호선": ("4호선", "안산과천선", "진접선"),
    "5호선": ("5호선",),
    "6호선": ("6호선",),
    "7호선": ("7호선", "도시철도 7호선"),
    "8호선": ("8호선", "수도권 광역철도 8호선"),
    "9호선": ("서울 도시철도 9호선", "수도권  도시철도 9호선"),
    "경강선": ("경강선",),
    "경의중앙선": ("경의중앙선", "경원선", "경부선"),
    "경춘선": ("경춘선", "경원선", "경의중앙선"),
    "공항철도": ("인천국제공항선",),
    "서해선": ("서해선",),
    "수인분당선": ("분당선", "수인선", "경원선", "안산과천선"),
    "신림선": ("수도권 경량도시철도 신림선",),
    "신분당선": ("신분당선",),
    "우이신설선": ("우이신설선",),
}

LINE_ORDER = [
    "1호선",
    "2호선",
    "3호선",
    "4호선",
    "5호선",
    "6호선",
    "7호선",
    "8호선",
    "9호선",
    "경의중앙선",
    "경춘선",
    "수인분당선",
    "신분당선",
    "공항철도",
    "경강선",
    "서해선",
    "우이신설선",
    "신림선",
    "GTX-A",
]
LINE_SORT_KEY = {name: index for index, name in enumerate(LINE_ORDER)}

NORMALIZED_NAME_ALIASES = {
    "지제": "평택지제",
    "신길온천": "능길",
    "뚝섬유원지": "자양",
    "세종왕릉": "세종대왕릉",
    "대모산": "대모산입구",
    "응암순환": "응암",
}

DISPLAY_NAME_OVERRIDES = {
    "지제": "평택지제",
    "신길온천": "능길",
    "뚝섬유원지": "자양",
    "세종왕릉": "세종대왕릉",
    "응암순환(상선)": "응암",
}

# The current official KRIC GTX-A workbook predates the northern section.
# These two opened stations are absent from it, so coordinates are supplemented
# from Rail.Blue and recorded explicitly in the build report.
GTX_COORDINATE_OVERRIDES = {
    "운정중앙": (37.716145841, 126.728092036),
    "킨텍스": (37.665204704, 126.748310364),
}
GTX_OVERRIDE_SOURCES = {
    "운정중앙": "https://rail.blue/railroad/logis/stationinfo.aspx?id=4821890",
    "킨텍스": "https://rail.blue/railroad/logis/stationinfo.aspx?id=4821891",
}

GTX_TRANSFER_LINES = {
    "대곡": {"일산선", "경의중앙선", "서해선"},
    "연신내": {"3호선", "6호선"},
    "서울": {"1호선", "4호선", "경부선", "경의중앙선", "인천국제공항선"},
    "수서": {"3호선", "분당선"},
    "성남": {"경강선"},
    "구성": {"분당선"},
}


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--kric", type=Path, default=DEFAULT_KRIC_PATH)
    parser.add_argument("--seoul", type=Path, default=DEFAULT_SEOUL_PATH)
    parser.add_argument("--gtx", type=Path, default=DEFAULT_GTX_PATH)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_PATH)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT_PATH)
    return parser.parse_args()


def read_sheet(path: Path, header_marker: str) -> list[dict[str, Any]]:
    if not path.is_file():
        raise FileNotFoundError(path)

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    header_index = next(
        index
        for index, row in enumerate(rows)
        if any(str(value).strip() == header_marker for value in row if value is not None)
    )
    headers = [str(value).strip() if value is not None else "" for value in rows[header_index]]

    records: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        if not any(value is not None and str(value).strip() for value in row):
            continue
        records.append({headers[index]: value for index, value in enumerate(row) if headers[index]})
    return records


def normalize_name(value: Any) -> str:
    name = unicodedata.normalize("NFC", str(value)).strip()
    name = re.sub(r"\([^)]*\)", "", name)
    name = re.sub(r"역$", "", name)
    name = re.sub(r"[\s.·ㆍ․-]", "", name)
    return NORMALIZED_NAME_ALIASES.get(name, name)


def identifier(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def coordinate(record: dict[str, Any], latitude_key: str, longitude_key: str) -> tuple[float, float]:
    return float(record[latitude_key]), float(record[longitude_key])


def haversine_meters(first: tuple[float, float], second: tuple[float, float]) -> float:
    latitude1, longitude1 = map(math.radians, first)
    latitude2, longitude2 = map(math.radians, second)
    latitude_delta = latitude2 - latitude1
    longitude_delta = longitude2 - longitude1
    value = (
        math.sin(latitude_delta / 2) ** 2
        + math.cos(latitude1) * math.cos(latitude2) * math.sin(longitude_delta / 2) ** 2
    )
    return 6_371_000 * 2 * math.atan2(math.sqrt(value), math.sqrt(1 - value))


def unique_coordinates(records: Iterable[dict[str, Any]]) -> list[tuple[float, float]]:
    coordinates = {
        tuple(round(value, 9) for value in coordinate(record, "역위도", "역경도"))
        for record in records
    }
    return sorted(coordinates)


def average_coordinate(coordinates: Iterable[tuple[float, float]]) -> tuple[float, float]:
    values = list(coordinates)
    if not values:
        raise ValueError("No coordinates to average")
    return (
        sum(value[0] for value in values) / len(values),
        sum(value[1] for value in values) / len(values),
    )


def select_kric_coordinate(
    seoul_record: dict[str, Any],
    kric_records: list[dict[str, Any]],
) -> tuple[float, float]:
    line_name = str(seoul_record["호선이름"])
    normalized_name = normalize_name(seoul_record["STATN_NM"])
    allowed_lines = LINE_ALIASES[line_name]

    candidates = [
        record
        for record in kric_records
        if record["normalizedName"] == normalized_name and record["노선명"] in allowed_lines
    ]
    exact_line_candidates = [record for record in candidates if record["노선명"] == line_name]
    if exact_line_candidates:
        candidates = exact_line_candidates

    if not candidates:
        # Some branch or transfer segments are recorded only under the other line
        # at the same physical station (for example 2호선 까치산 under 5호선).
        candidates = [
            record for record in kric_records if record["normalizedName"] == normalized_name
        ]

    if not candidates:
        raise ValueError(
            f"No coordinate match: {line_name} {seoul_record['STATN_NM']} "
            f"({seoul_record['STATN_ID']})"
        )

    coordinates = unique_coordinates(candidates)
    if any(
        haversine_meters(first, second) > CLUSTER_DISTANCE_METERS
        for index, first in enumerate(coordinates)
        for second in coordinates[index + 1 :]
    ):
        raise ValueError(
            f"Ambiguous coordinate match: {line_name} {seoul_record['STATN_NM']} "
            f"has {len(coordinates)} distant candidates"
        )
    return average_coordinate(coordinates)


def select_gtx_coordinate(
    seoul_record: dict[str, Any],
    kric_records: list[dict[str, Any]],
    gtx_coordinates: dict[str, tuple[float, float]],
) -> tuple[float, float]:
    normalized_name = normalize_name(seoul_record["STATN_NM"])
    transfer_lines = GTX_TRANSFER_LINES.get(normalized_name)
    candidates = [
        record
        for record in kric_records
        if record["normalizedName"] == normalized_name
        and (transfer_lines is None or record["노선명"] in transfer_lines)
    ]
    if candidates:
        coordinates = unique_coordinates(candidates)
        if all(
            haversine_meters(first, second) <= CLUSTER_DISTANCE_METERS
            for index, first in enumerate(coordinates)
            for second in coordinates[index + 1 :]
        ):
            return average_coordinate(coordinates)

    if normalized_name in gtx_coordinates:
        return gtx_coordinates[normalized_name]
    if normalized_name in GTX_COORDINATE_OVERRIDES:
        return GTX_COORDINATE_OVERRIDES[normalized_name]
    raise ValueError(f"No GTX-A coordinate match: {seoul_record['STATN_NM']}")


def connected_components(
    normalized_name: str,
    records: list[dict[str, Any]],
) -> list[list[dict[str, Any]]]:
    if normalized_name in DISTINCT_SAME_NAME_LINES:
        components: list[list[dict[str, Any]]] = []
        assigned_lines: set[str] = set()
        for line_group in DISTINCT_SAME_NAME_LINES[normalized_name]:
            component = [row for row in records if str(row["호선이름"]) in line_group]
            if component:
                components.append(component)
                assigned_lines.update(str(row["호선이름"]) for row in component)
        unassigned = [row for row in records if str(row["호선이름"]) not in assigned_lines]
        if unassigned:
            raise ValueError(f"Unassigned distinct-station lines: {normalized_name}")
        return components

    if normalized_name in FORCED_TRANSFER_NAMES:
        return [records]

    remaining = set(range(len(records)))
    components: list[list[dict[str, Any]]] = []

    while remaining:
        pending = [remaining.pop()]
        component_indexes: set[int] = set()
        while pending:
            current = pending.pop()
            if current in component_indexes:
                continue
            component_indexes.add(current)
            neighbors = {
                index
                for index in remaining
                if haversine_meters(records[current]["coordinate"], records[index]["coordinate"])
                <= CLUSTER_DISTANCE_METERS
            }
            remaining.difference_update(neighbors)
            pending.extend(neighbors)
        components.append([records[index] for index in sorted(component_indexes)])
    return components


def choose_api_name(records: list[dict[str, Any]]) -> str:
    counts = Counter(str(record["STATN_NM"]).strip() for record in records)
    return min(counts, key=lambda value: (-counts[value], len(value), value))


def build_station(component: list[dict[str, Any]]) -> dict[str, Any]:
    api_name = choose_api_name(component)
    display_names = [DISPLAY_NAME_OVERRIDES.get(str(row["STATN_NM"]), str(row["STATN_NM"])) for row in component]
    display_name = min(display_names, key=lambda value: (len(value), value))
    normalized_name = normalize_name(component[0]["STATN_NM"])
    if any(normalize_name(row["STATN_NM"]) != normalized_name for row in component):
        raise ValueError("A physical station component contains multiple normalized names")
    coordinates = [row["coordinate"] for row in component]
    latitude, longitude = PHYSICAL_STATION_COORDINATE_OVERRIDES.get(
        normalized_name,
        average_coordinate(coordinates),
    )
    line_names = sorted(
        {str(row["호선이름"]) for row in component},
        key=lambda value: (LINE_SORT_KEY.get(value, 999), value),
    )

    station_ids: dict[str, str] = {}
    subway_ids: dict[str, str] = {}
    for row in component:
        line_name = str(row["호선이름"])
        station_id = identifier(row["STATN_ID"])
        subway_id = identifier(row["SUBWAY_ID"])
        if line_name in station_ids and station_ids[line_name] != station_id:
            raise ValueError(f"Multiple station IDs for one physical line: {display_name} {line_name}")
        station_ids[line_name] = station_id
        subway_ids[line_name] = subway_id

    stable_station_id = min(station_ids.values())
    return {
        "id": f"seoul-{stable_station_id}",
        "name": display_name,
        "apiName": api_name,
        "latitude": round(latitude, 7),
        "longitude": round(longitude, 7),
        "lineNames": line_names,
        "seoulStationIDs": dict(sorted(station_ids.items(), key=lambda item: LINE_SORT_KEY.get(item[0], 999))),
        "subwayIDs": dict(sorted(subway_ids.items(), key=lambda item: LINE_SORT_KEY.get(item[0], 999))),
    }


def validate_stations(stations: list[dict[str, Any]], source_rows: int) -> dict[str, Any]:
    ids = [station["id"] for station in stations]
    if len(ids) != len(set(ids)):
        raise ValueError("Duplicate station IDs in output")

    represented_rows = sum(len(station["lineNames"]) for station in stations)
    if represented_rows != source_rows:
        raise ValueError(f"Expected {source_rows} station-line rows, represented {represented_rows}")

    for station in stations:
        if not 33.0 <= station["latitude"] <= 39.0:
            raise ValueError(f"Latitude outside Korea: {station}")
        if not 124.0 <= station["longitude"] <= 132.0:
            raise ValueError(f"Longitude outside Korea: {station}")
        if set(station["lineNames"]) != set(station["seoulStationIDs"]):
            raise ValueError(f"Line and station ID mismatch: {station['name']}")

    name_counts = Counter(station["name"] for station in stations)
    duplicate_physical_names = sorted(name for name, count in name_counts.items() if count > 1)
    line_counts = Counter(
        line_name for station in stations for line_name in station["lineNames"]
    )
    return {
        "stationCount": len(stations),
        "sourceStationLineCount": source_rows,
        "lineCounts": dict(sorted(line_counts.items(), key=lambda item: LINE_SORT_KEY.get(item[0], 999))),
        "sameNameDifferentLocation": duplicate_physical_names,
        "gtxCoordinateOverrides": GTX_OVERRIDE_SOURCES,
        "physicalStationCoordinateOverrides": {
            name: {
                "latitude": coordinate[0],
                "longitude": coordinate[1],
                "reason": "KRIC 경원선 좌표 채택; 동일 환승역의 4호선 좌표 이상치 보정",
            }
            for name, coordinate in PHYSICAL_STATION_COORDINATE_OVERRIDES.items()
        },
        "distinctSameNameStations": sorted(DISTINCT_SAME_NAME_LINES),
    }


def main() -> None:
    arguments = parse_arguments()
    kric_records = read_sheet(arguments.kric, "역번호")
    seoul_records = read_sheet(arguments.seoul, "SUBWAY_ID")
    gtx_records = read_sheet(arguments.gtx, "No")

    for record in kric_records:
        record["normalizedName"] = normalize_name(record["역사명"])

    gtx_coordinates: dict[str, tuple[float, float]] = {}
    for record in gtx_records:
        longitude = record.get("역 위치(경도)")
        latitude = record.get("역 위치(위도)")
        if latitude is None or longitude is None:
            continue
        gtx_coordinates[normalize_name(record["역명(한글)"])] = (float(latitude), float(longitude))

    matched_rows: list[dict[str, Any]] = []
    for record in seoul_records:
        line_name = str(record["호선이름"])
        record["coordinate"] = (
            select_gtx_coordinate(record, kric_records, gtx_coordinates)
            if line_name == "GTX-A"
            else select_kric_coordinate(record, kric_records)
        )
        record["normalizedName"] = normalize_name(record["STATN_NM"])
        matched_rows.append(record)

    grouped_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in matched_rows:
        grouped_rows[record["normalizedName"]].append(record)

    stations: list[dict[str, Any]] = []
    for normalized_name in sorted(grouped_rows):
        for component in connected_components(normalized_name, grouped_rows[normalized_name]):
            stations.append(build_station(component))

    stations.sort(key=lambda station: (station["name"], station["latitude"], station["longitude"]))
    report = validate_stations(stations, len(seoul_records))
    report["sources"] = {
        "kric": arguments.kric.name,
        "seoulRealtimeSupport": arguments.seoul.name,
        "kricGtxA": arguments.gtx.name,
    }
    report["clusterDistanceMeters"] = CLUSTER_DISTANCE_METERS

    arguments.output.parent.mkdir(parents=True, exist_ok=True)
    arguments.report.parent.mkdir(parents=True, exist_ok=True)
    arguments.output.write_text(json.dumps(stations, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    arguments.report.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print("Station bundle build completed")
    print(f"- Seoul station-line rows: {len(seoul_records)}")
    print(f"- Physical stations: {len(stations)}")
    print(f"- Same-name distinct stations: {len(report['sameNameDifferentLocation'])}")
    print(f"- Output: {arguments.output.relative_to(PROJECT_ROOT)}")
    print(f"- Report: {arguments.report.relative_to(PROJECT_ROOT)}")


if __name__ == "__main__":
    main()
